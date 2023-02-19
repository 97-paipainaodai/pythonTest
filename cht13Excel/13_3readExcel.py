import openpyxl
wb = openpyxl.load_workbook('D:\yss\TEST\cht13Excel\example.xlsx')
print(type(wb))
print(wb.sheetnames)   #the workbook's sheets names

sheet1 = wb['Sheet3']   #Get a sheet from the workbook
print(sheet1)
print(sheet1.title)    #Get the sheet's title as a string

sheet2 = wb.active   #Get the active sheet(only sheet1 has data)
print(sheet2)

sheet = wb['Sheet1']  #Get a sheet from the workbook
print(sheet['A1'])   #Get a cell from the sheet
print(sheet['A1'].value)   #Get the value from the cell

c = sheet['B1']  #Get another cell from the sheet
print(c.value)

#Get the row, column, and value from the cell
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))  #Row 1, Column 2 is Apples
print('Cell %s is %s' % (c.coordinate, c.value))  #Cell B1 is Apples

#Use number 1,2,3... to express row and column
print(sheet.cell(row=1, column=2))  #<Cell 'Sheet1'.B1>
rowMax = sheet.max_row
colMax = sheet.max_column
print('rowMax = ', rowMax)
print('colMax = ', colMax)
for i in range(1,8,2): #Go through every other row
    print(i, sheet.cell(row=i, column=2).value)

#Translate column number to letter
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))  #A
print(get_column_letter(2))  #B
print(get_column_letter(27))  #AA
print(get_column_letter(sheet.max_column))

print(column_index_from_string('A'))  #Get A's number

print(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

print(list(sheet.columns)[1])  #Get second column's cells
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
